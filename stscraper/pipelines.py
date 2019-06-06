# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html
import openpyxl
#from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

class StscraperPipeline(object):
    def open_spider(self, spider):
        self.file = openpyxl.load_workbook("data.xlsx")
        self.sheet = self.file["Sheet1"]

    def close_spider(self, spider):
        self.file.save("data.xlsx")
        # self.file.close()

    @classmethod
    def formatting(self, workbook, sheetname):
        wb = openpyxl.load_workbook(workbook)
        sheet = wb[sheetname]

        for col in sheet.columns:
            max_length = 0
            column = col[0].column

            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[get_column_letter(column)].width = adjusted_width
        wb.save(workbook)

    def process_item(self, item, spider):

        next_row = 1
        if self.sheet.max_row >= 1:
            next_row = self.sheet.max_row + 2
        else:
            next_row = 1

        j = 1
        for key, value in item.items():
            self.sheet.cell(row=next_row, column=j).value = str(value)
            # self.sheet.cell(row=next_row + 1, column=j).value = str(value)
            j += 1
        self.file.save("data.xlsx")
        self.formatting("data.xlsx", "Sheet1")

        return item


